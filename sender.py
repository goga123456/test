import email
import email.mime.application
import smtplib
import ssl
from datetime import datetime, time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook


filename = 'example.xlsx'


async def send_email():
    msg = MIMEMultipart("alternative")
    fromaddr = "bukanov1234@mail.ru"
    mypass = "6bUc5jT7is5Yvz4pYHLf"
    toaddr = "rezume_BOT@beeline.uz"
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "Отправитель: Клон Telegram bot"

    now = datetime.now()
    response_date = now.strftime("%d.%m.%Y")

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
    </head>
    <body>        
    <h1>Отчёт за: {response_date} </h1>      
    </body>
    </html>
    """
    text = bs(html, "html.parser").text
    msg.attach(MIMEText(text, 'plain'))
    msg.attach(MIMEText(html, 'html', 'utf-8'))
    fp = open(filename, 'rb')
    att = email.mime.application.MIMEApplication(fp.read(), _subtype="xlsx")
    fp.close()
    att.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(att)

    server = smtplib.SMTP_SSL('smtp.mail.ru:465')
    ssl.SSLContext(ssl.PROTOCOL_TLS)
    server.login(msg['From'], mypass)
    text = msg.as_string()
    server.sendmail(msg['From'], msg['To'], text)
    server.quit()

    print("Successfully")
    await clear_sheet()


async def clear_sheet():
    wb = load_workbook(filename=filename)
    ws = wb['Лист1']
    nb_row = ws.max_row
    ws.delete_rows(2, nb_row)
    ws = wb['Лист2']
    ws['A2'] = 0
    ws['B2'] = 0
    ws['C2'] = 0
    ws['D2'] = 0
    wb.save(filename)
